import pandas as pd
from openpyxl import load_workbook
from twilio.rest import Client
from dotenv import load_dotenv
import os
import time
from flask import Flask, request, jsonify, send_file
from twilio.twiml.voice_response import VoiceResponse
import threading
from pyngrok import ngrok
from pathlib import Path
import openai
import requests
from datetime import datetime
import shutil
from pydub import AudioSegment
from queue import Queue
import re
import ast
# komentoffftest
# Flaskアプリケーションの作成
app = Flask(__name__)

# ファイルから環境変数を読み込む
print("環境変数を.envファイルから読み込みます")
load_dotenv()

# グローバル変数
audio_file_path = 'audio.mp3'
output_file_path = 'output.mp3'
account_sid = os.getenv('TWILIO_SID')
auth_token = os.getenv('TWILIO_AUTH_TOKEN')
from_number = os.getenv('TWILIO_PHONE_NUMBER')
my_info="株式会社テスト、業務改革室の竹内です。"
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
first_or = True  # 初期値としてTrueを設定
vacant_enough=False
call_end=False
roomnumber_enough=False
building_enough=False
conversation_history=[]
room_details=[]
vacant_rooms=[]
building_data = []
missing_info=[]
room_notenough=[]

# 現在の日付を取得
current_date = datetime.now().strftime('%Y/%m/%d')

# Twilioの設定
print("Twilioクライアントを初期化します")
client = Client(account_sid, auth_token)
print("Twilioクライアントが初期化されました")

# Excelファイルのパスを設定
print("Excelファイルのパスを設定します")
excel_file_path = 'C:\\Users\\ntake\\my_project\\bukkaku_call_automation\\デモ用ビルデータサンプル.xlsx'
print(f"Excelファイルパス: {excel_file_path}")

# コピーするExcelファイルのパス
copied_file_path = f'{excel_file_path.split(".")[0]}_copy.xlsx'
shutil.copyfile(excel_file_path, copied_file_path)

# 会話履歴のファイルパスを設定
conversation_history_file_path = 'C:\\Users\\ntake\\my_project\\bukkaku_call_automation\\conversation_history.txt'

# Excelファイルを読み込み
print("Excelファイルを読み込みます")
df = pd.read_excel(copied_file_path, dtype={'電話番号': str, 'ビルコード（7ケタ）': str,"ビル名称":str})

# openpyxlを使用して同じExcelファイルを開く
wb = load_workbook(copied_file_path)

# Twilioの環境変数が正しく設定されているか確認
if not account_sid or not auth_token or not from_number:
    print("Twilioの環境変数が正しく設定されていません。")
    raise ValueError("Twilioの環境変数が正しく設定されていません。")

# 電話番号の冒頭の0を+81に変換する関数
def convert_phone_number(phone_number):
    print(f"電話番号を変換します: {phone_number}")
    phone_number = str(phone_number)  # 文字列に変換
    if phone_number.startswith('0'):
        converted = '+81' + phone_number[1:]
        print(f"変換後の電話番号: {converted}")
        return converted
    print("電話番号は変換の必要がありません")
    return phone_number

# 会話履歴から空き部屋の部屋番号を抽出する関数
def extract_vacant_rooms_from_conversation(conversation_history):
    """
    conversation_history.txtから会話履歴を読み込み、
    ChatGPTで分析して空き部屋の部屋番号をリストとして抽出する関数。
    
    :param file_path: conversation_history.txt のファイルパス
    :return: 空き部屋の部屋番号をリストとして返す
    """
    try:
        # ChatGPT APIを利用して、空き部屋の部屋番号を抽出するプロンプトを作成
        prompt = f"以下の会話履歴から、空き部屋として紹介されている部屋番号をすべてリストアップしてください。\n\n{conversation_history}"

        # ChatGPT APIを使用して会話の内容を分析
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role":"system","content":"あなたは不動産仲介会社の事務員です。"},
                {"role":"user","content":f'{prompt}'}
            ]
        )

        # ChatGPTの応答から空き部屋の部屋番号をリストとして抽出
        vacant_rooms_text = response.choices[0].message.content

        # 空き部屋の部屋番号をリストに変換
        vacant_rooms_list = [room.strip() for room in vacant_rooms_text.split('\n') if room]

        return vacant_rooms_list

    except Exception as e:
        print(f"会話履歴からの部屋番号取得時にエラーが発生しました: {e}")
        return []

# 会話履歴から空き部屋の各情報（結果、部屋番号）が十分に収集されているか判定する関数
def judgement_enough(file_path, conversation_history,model="gpt-4o-mini"):
    element="部屋番号"
    # 会話履歴をテキストファイルに保存
    save_conversation_history(conversation_history_file_path, conversation_history)
    # move_file_to_date_folder(conversation_history_file_path, f"{company_name}_{building_name}")

    # 会話履歴を読み込む
    with open(file_path, 'r', encoding='utf-8') as file:
        conversation_history = file.read()
    print(conversation_history)
    
    if element == "部屋番号":

        response = openai.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": f"あなたは不動産の専門家です。"},
                {"role": "user", "content": f'以下の会話履歴から、{element}に関する情報が十分に収集されているかどうかを判定し、「十分」または「不十分」の結果のみを出力してください。このとき、システムの、「〇〇号室の他に空いている部屋は無いか」といった旨の質問に対して否定的な回答を得た際には「十分」としてよい。\n{conversation_history}'}
            ]
        )
    else:
        response = openai.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": f"あなたは不動産の専門家です。"},
                {"role": "user", "content": f'以下の会話履歴から、{element}に関する情報が十分に収集されているかどうかを判定し、「十分」または「不十分」の結果のみを出力してください。\n{conversation_history}'}
            ]
        )
    print(response.choices[0].message.content)
    analysis_result = response.choices[0].message.content

    # "不十分" や "不足している" のような否定的な表現が含まれていないかも確認する
    if any(keyword in analysis_result for keyword in ["十分", "complete", "全ての情報", "sufficient"]) and \
       not any(neg in analysis_result for neg in ["不十分", "不足", "足りない", "missing", "incomplete"]):
        return True
    else:
        return False
  
# 音声ファイルを結合、またはrecording1が空の場合はrecording2のコピーを作成する関数
def concatenate_or_copy_recordings(recording1, recording2, output_file_path):
    """
    2つの録音ファイルを結合、またはrecording2が空の場合はrecording1のコピーを作成する関数
    
    :param recording1: 1つ目の録音ファイルのパス（空の場合はrecording1のコピーを作成）
    :param recording2: 2つ目の録音ファイルのパス
    :param output_file_path: 出力するファイルのパス
    """
    try:
        if not recording1:  # recording2が空の場合
            shutil.copyfile(recording2, output_file_path)
            print(f"{recording2}のコピーが作成されました: {output_file_path}")
        elif not recording2:  # recording1が空の場合
            shutil.copyfile(recording1, output_file_path)
            print(f"{recording1}のコピーが作成されました: {output_file_path}")
        elif recording1==output_file_path:
            output_file_path_backup=output_file_path.split(".")[0]+"_backup.mp3"
            shutil.copyfile(output_file_path, output_file_path_backup)
            # 各音声ファイルをロード
            audio1 = AudioSegment.from_file(output_file_path_backup)
            audio2 = AudioSegment.from_file(recording2)
            
            # 音声ファイルを結合
            combined_audio = audio1 + audio2
            
            # 結合した音声ファイルを保存
            combined_audio.export(output_file_path, format="mp3")
            print(f"音声ファイルが結合され、保存されました: {output_file_path}")
        else:
            # 各音声ファイルをロード
            audio1 = AudioSegment.from_file(recording1)
            audio2 = AudioSegment.from_file(recording2)
            
            # 音声ファイルを結合
            combined_audio = audio1 + audio2
            
            # 結合した音声ファイルを保存
            combined_audio.export(output_file_path, format="mp3")
            print(f"音声ファイルが結合され、保存されました: {output_file_path}")
    except Exception as e:
        print(f"音声ファイルの処理中にエラーが発生しました: {e}")

# 音声をテキスト化する関数
def transcribe_audio(audio_file_path: str, model: str = "whisper-1") -> str:
    print(f"音声ファイルをテキスト化しています: {audio_file_path}")
    with open(audio_file_path, "rb") as audio_file:
        transcript = openai.audio.transcriptions.create(
            model=model,
            file=audio_file,
            language="ja"  # 言語を日本語に指定
        )
    print("音声ファイルのテキスト化が完了しました")
    return transcript.text

# ChatGPT 4o miniを使用して応答を生成する関数
def generate_response(stance: str, conversation_history: list, model: str = "gpt-4o-mini") -> str:
    print("ChatGPT-4.0-miniを使用して応答を生成しています")
    response = openai.chat.completions.create(
      model=model,
      messages=[
        {"role": "system", "content": "あなたは不動産仲介会社の営業マンで、不動産管理会社に対して、管理会社が管理する物件の各部屋について空き状況を確認する目的で電話をかけています。"},
        *conversation_history,
        {"role": "user", "content": "話し相手は電話口の不動産仲介会社です。相手の発言に対して丁寧に、極力短い文字数で対応してください。あなたが話す文言だけを生成してくれればよいです。"+stance},
      ]
    )
    print("応答が生成されました")
    return response.choices[0].message.content

# テキストを音声ファイル化する関数
def text_to_speech(text: str, output_file_path: str, model: str = "tts-1", voice: str = "alloy"):
    print(f"テキストを音声ファイルに変換しています: {text}")
    speech_file_path = Path(output_file_path)
    
    # OpenAI APIを使用して音声を生成
    response = openai.audio.speech.create(
        model=model,
        voice=voice,
        input=text
    )
    with open(speech_file_path, 'wb') as f:
        f.write(response.content)
        print(f"音声ファイルが生成されました: {speech_file_path}")

# 会話履歴をテキストファイルに保存する関数
def save_conversation_history(file_path: str, conversation_history: list):
    try:
        print(f"会話履歴をファイルに保存します: {file_path}")
        with open(file_path, 'w', encoding='utf-8') as file:
            for entry in conversation_history:
                # roleとcontentのペアを取り出して、ファイルに書き込む
                role = entry.get('role', 'unknown')  # roleがない場合は'unknown'を使用
                content = entry.get('content', '')  # contentがない場合は空文字列を使用
                file.write(f"{role}: {content}\n")
        print(f"会話履歴が {file_path} に保存されました")
    except Exception as e:
        print(f"会話履歴の保存中にエラーが発生しました: {e}")

# 空のビルデータを格納したリストに空き部屋情報を追加する関数
def add_vacant_rooms_to_building(building_data,building_name, vacant_room_numbers):
    """
    vacant_rooms から部屋番号を抽出し、building_name ごとに対応する辞書を building_data リストに格納する関数
    
    :param building_name: ビル名
    :param vacant_rooms: 部屋番号が記載されたリスト
    :return: 更新された building_data
    """
    print(f"[INFO] Building name: {building_name}")
    print(f"[INFO] Vacant rooms before processing: {vacant_room_numbers}")

    # 部屋番号ごとの情報を作成
    room_info_list = [{'部屋番号': room, '坪数': '', '坪単価': '', '保証金': '', '入居時期': '', '備考': ''} for room in vacant_room_numbers]
    
    # building_data に辞書を追加
    building_data.append({building_name: room_info_list})
    
    return building_data

# ビルデータを格納したリスト内の辞書を、部屋番号数字昇順→アルファベット昇順でソートする関数
def sort_vacant_rooms_in_building(building_data):
    """
    building_data の中の各ビルの部屋番号を数字昇順→アルファベットの昇順でソートし、
    その辞書の順番を入れ替える関数。
    
    :param building_data: 部屋情報を含むビルごとの辞書のリスト
    :return: ソート後の building_data
    """
    
    def room_sort_key(room):
        # 部屋番号の数字部分とアルファベット部分を抽出してタプルを作成
        room_number = room['部屋番号']
        number_part = ''.join(filter(str.isdigit, room_number))
        alpha_part = ''.join(filter(str.isalpha, room_number))
        return (int(number_part) if number_part else 0, alpha_part)
    
    sorted_building_data = []
    
    for building in building_data:
        for building_name, rooms in building.items():
            # 部屋番号をソート
            sorted_rooms = sorted(rooms, key=room_sort_key)
            sorted_building_data.append({building_name: sorted_rooms})
    
    return sorted_building_data

# 指定文言で始まる全ての列を削除する関数（いまは空室情報固定）
def remove_vacant_columns(df):
    """
    DataFrameのカラム名が「空室情報」で始まる全ての列を削除する関数

    Parameters:
    - df: 操作対象のDataFrame

    Returns:
    - 更新後のDataFrame
    """
    # カラム名が「空室情報」で始まる列を見つけて削除
    cols_to_remove = [col for col in df.columns if col.startswith("空室情報")]
    print(f"削除対象のカラム: {cols_to_remove}")
    df.drop(columns=cols_to_remove, inplace=True)
    print("指定されたカラムを削除しました。")

    return df

# ビルデータ内情報をエクセルに書き込む関数
def insert_vacant_info_to_columns(building_data, df):
    """
    building_dataを元に、ビル名称が一致する行に対してH列から順に空室情報を挿入する関数

    Parameters:
    - building_data: 空室情報が格納されたリスト（辞書のリスト）
    - df: 操作対象のDataFrame

    Returns:
    - 更新後のDataFrame
    """
    for building in building_data:
        for building_name, room_info_list in building.items():
            # 「ビル名称」が一致する行を探す
            matching_rows = df[df['ビル名称'] == building_name]
            
            for index, row in matching_rows.iterrows():
                print(f"行 {index+1}: ビル名称は {building_name} です。")
                
                # H列から順に空室情報を格納していく
                for i, room_info in enumerate(room_info_list):
                    col_index = 7 + i  # H列は8番目のカラムでindexは7、I列は9番目のカラムでindexは8
                    col_name = df.columns[col_index] if col_index < len(df.columns) else None
                    
                    # 列が存在しない、またはカラム名が「空室情報」で始まっていない場合
                    if col_name is None or not col_name.startswith("空室情報"):
                        new_col_name = f"空室情報{i+1}"
                        df.insert(col_index, new_col_name, "")
                        col_name = new_col_name
                        print(f"新しいカラム '{col_name}' を挿入しました。")
                    
                    # 辞書内の情報をテキストにまとめる
                    room_info_text = "\n".join([f"{key}: {value}" for key, value in room_info.items()])
                    
                    # 指定列に情報を格納
                    df.at[index, col_name] = room_info_text
                    print(f"行 {index+1}, 列 {col_name} を更新しました: {room_info_text}")
    
    return df

# 会話履歴から部屋番号を抽出する関数
def extract_room_numbers(conversation_history):

    print(f"[INFO] Extracting room numbers from: {conversation_history}")

    try:
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "あなたは、Pythonのリスト形式でデータを出力するAIです。"},
                {"role": "user", "content": f"以下の文に対して、Pythonのリスト形式で部屋番号を出力してください。リストは部屋番号のみを含み、引用符で囲まれた文字列形式で出力してください。\n例: ['101', '102', '103']\n{conversation_history}"}
            ]
        )
        print(f"[INFO] GPT-4 response received: {response}")
    except Exception as e:
        print(f"[ERROR] Failed to extract room numbers: {e}")
        return []

    # GPT-4からの応答をリスト形式で出力
    try:
        # 応答から部屋番号のリストを抽出
        content = response.choices[0].message.content
        print(f"[INFO] Extracted content: {content}")
        
        # 正規表現を使用してリストを抽出
        match = re.search(r'\[([^\[\]]+)\]', content)
        if match:
            room_numbers = match.group(1).split(',')
            room_numbers = [num.strip().strip('"').strip("'") for num in room_numbers]
            print(f"[INFO] Extracted room numbers: {room_numbers}")
            return room_numbers
        else:
            print("[ERROR] Failed to parse room numbers from GPT-4 response")
            return []
    except Exception as e:
        print(f"[ERROR] Failed to process GPT-4 response: {e}")
        return []

# 会話履歴から部屋の詳細情報を抽出し、辞書を生成する関数
def extract_room_details(missing_info,conversation_history):

    try:
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "あなたは、Pythonの辞書形式でデータを出力するAIです。"},
                {"role": "user", "content": f"以下の文に対して、Pythonの辞書形式で{missing_info}のそれぞれのキーに対応する値を抽出し、正しい位置に格納してください。各値は引用符で囲まれた文字列形式で格納してください。\n{conversation_history}"}
            ]
        )
        print(f"[INFO] GPT-4 response received: {response}")
    except Exception as e:
        print(f"[ERROR] Failed to extract room numbers: {e}")
        return []

    # GPT-4からの応答をリスト形式で出力
    try:
        # 応答から部屋番号のリストを抽出
        content = response.choices[0].message.content
        print(f"[INFO] Extracted content: {content}")
        return content
        
    except Exception as e:
        print(f"[ERROR] Failed to process GPT-4 response: {e}")
        return []

# ファイル名を変えて、今日の日付のフォルダに移動フォルダに格納する関数
def move_file_to_date_folder(filepath, new_name):
    global room_numbers
    # 日付をYYYYMMDD形式で取得
    date_str = datetime.now().strftime('%Y%m%d')

    # 新しいフォルダパスを作成
    directory = os.path.dirname(filepath)
    date_folder = os.path.join(directory, date_str)
    company_folder=os.path.join(date_folder,company_name)

    # フォルダが存在しない場合は作成
    if not os.path.exists(date_folder):
        os.makedirs(date_folder)
        print(f"フォルダを作成しました: {date_folder}")
    else:
        print(f"フォルダが既に存在します: {date_folder}")
    if not os.path.exists(company_folder):
        os.makedirs(company_folder)
        print(f"フォルダを作成しました: {company_folder}")
    else:
        print(f"フォルダが既に存在します: {company_folder}")

    # 元のファイルの拡張子を取得
    _, file_extension = os.path.splitext(filepath)

    # 新しいファイル名と拡張子を組み合わせる
    new_filename = new_name + file_extension
    new_filepath = os.path.join(company_folder, new_filename)

    # 同じ名前のファイルが存在する場合、(1), (2), ... のように番号を付ける
    counter = 1
    while os.path.exists(new_filepath):
        new_filename_with_suffix = f"{new_name}({counter}){file_extension}"
        new_filepath = os.path.join(company_folder, new_filename_with_suffix)
        counter += 1

    # ファイルを新しい場所に移動
    os.rename(filepath, new_filepath)
    print(f"ファイルを {filepath} から {new_filepath} に移動しました。")

    return new_filepath

# {}を含む文字列から辞書を生成する関数
def dict_from_str(room_detail):
    # 正規表現を使用して、{}で囲まれた部分を抽出
    dict_str_match = re.search(r'\{.*\}', room_detail, re.DOTALL)

    if dict_str_match:
        dict_str = dict_str_match.group()

        # インデントと不要な改行を削除
        dict_str = re.sub(r'^[ \t]+', '', dict_str, flags=re.MULTILINE)

        # シングルクォートをダブルクォートに変換 (JSON風にするため)
        dict_str = dict_str.replace("'", '"')

        # Noneをnullに変換 (JSON形式に準拠)
        dict_str = dict_str.replace('None', 'null')

        try:
            # 辞書として変換
            room_detail_dict = ast.literal_eval(dict_str)
            return room_detail_dict
        except Exception as e:
            print(f"辞書への変換に失敗しました: {e}")
            return None
    else:
        print("辞書が見つかりませんでした")
        return None

# TwiMLエンドポイント(メインのエンドポイント)
@app.route('/twiml', methods=['POST'])
def twiml():
    print("TwiMLエンドポイントが呼び出されました")
    
    # TwiMLレスポンスを生成
    response = VoiceResponse()
    
    # 挨拶用の音声ファイルを再生
    response.play(public_url + '/serve_audio')  # 音声ファイルのURL
    print("音声ファイルを再生しました")
    
    # 再生が完了したら録音を開始するようにRedirect
    response.redirect('/start-recording')
    print("Redirect to /start-recording")
    
    print("TwiMLレスポンスが生成されました")

    return str(response)

# Flaskルート: システム発話用の音声ファイルを提供するためのエンドポイント
@app.route('/serve_audio')
def serve_audio():
    print("システム発話用の音声ファイルを提供しています")
    try:
        print("システム発話用の音声ファイルを提供しています")
        return send_file(audio_file_path, mimetype='audio/mpeg')
    except Exception as e:
        print(f"システム発話用の音声ファイルの提供中にエラーが発生しました: {e}")
        return jsonify(error=str(e)), 500

# 録音を開始するエンドポイント
@app.route('/start-recording', methods=['POST'])
def start_recording():
    print("録音を開始します")
    response = VoiceResponse()
    response.record(max_length=60, action='/handle-recording', method="POST", transcribe=False)
    print("録音を開始しました")
    return str(response)

# 録音が完了した後のエンドポイント
@app.route('/handle-recording', methods=['POST'])
def handle_recording():

    global conversation_history
    global vacant_enough
    global vacant_room_numbers
    global details_enough
    global building_data
    global building_name
    global missing_info
    global room_notenough
    global company_name
    global building_names
    global room_details
    global building_number
    global roomnumber_enough
    global building_enough
    call_end=False
        
    print("録音処理を開始します")
    
    recording_url = request.form['RecordingUrl']
    print(f"録音URLを受信しました: {recording_url}")
   
    time.sleep(3)  # 3秒待機
    print("3秒待機しました。録音ファイルをダウンロードします。")

    try:
        response = requests.get(recording_url, auth=(account_sid, auth_token))
        response.raise_for_status()
        print("録音ファイルのダウンロードが成功しました。")

        with open(audio_file_path, 'wb') as audio_file:
            audio_file.write(response.content)
            print(f"録音ファイルを保存しました: {audio_file_path}")
        # 音声ファイルを結合        
        concatenate_or_copy_recordings(output_file_path, audio_file_path, output_file_path)

        # 録音ファイルをテキスト化
        transcription = transcribe_audio(audio_file_path)
        print(f"テキスト化された内容: {transcription}")

        # 会話履歴に追加
        conversation_history.append({"role": "user", "content": transcription})
        print(f"会話履歴に追加しました：{conversation_history}")
        # conversation_txt_file_path=f"{conversation_history_file_path.split(".")[0]}_{company_name}_{building_name}.txt"
        
        # 対象ビルの空き部屋の部屋番号が十分に収集されているかどうかを判定
        if building_enough!=True:
            building_enough=judgement_enough(conversation_history_file_path, conversation_history)

        # 対象ビルの空き部屋の部屋番号が十分に収集されている場合
        if building_enough==True and vacant_enough==True and roomnumber_enough!=True:
            # 会話履歴から空き部屋の部屋番号を抽出
            print(f"[INFO] Vacant rooms before processing: {conversation_history}")
            vacant_room_numbers = extract_room_numbers(conversation_history)
            print(f"[INFO] Extracted vacant room numbers: {vacant_room_numbers}")
            print("空き部屋の部屋番号が十分に収集されています")

            # ビルデータ（部屋番号のみ）の作成
            building_data = add_vacant_rooms_to_building(building_data,building_name, vacant_room_numbers)
            print(f"[INFO] Updated building data: {building_data}")
            print(building_data)
            print(f"ビルデータに空室情報を追加しました: {building_data}")
            # ビルデータ内の情報を部屋番号でソート
            building_data = sort_vacant_rooms_in_building(building_data)

            roomnumber_enough=True
        
        # 対象ビルの空き部屋の部屋番号が十分に収集されていない場合
        elif roomnumber_enough!=True:
            # 会話履歴から空き部屋の部屋番号を抽出
            print(f"[INFO] Vacant rooms before processing: {conversation_history}")
            vacant_room_numbers = extract_room_numbers(conversation_history)
            print(f"[INFO] Extracted vacant room numbers: {vacant_room_numbers}")
            print("空き部屋の部屋番号が不十分です")
            # 空き部屋の部屋番号収集を続けるための応答生成
            stance=f"現状確認できている空き部屋の部屋番号は以下の通りです。{vacant_room_numbers}ここで空き確認が取れている部屋番号を相手に伝え、他にも空いている部屋が無いか尋ねてください。"
            vacant_enough=True

        # 各部屋の詳細情報収集
        if roomnumber_enough==True:
            print("部屋番号の情報が十分に収集されたため、部屋の詳細情報を抽出します。")
            
            if missing_info:
                room_details=extract_room_details(missing_info,conversation_history)
                room_details=dict_from_str(room_details)
                print(f"部屋の詳細情報:{room_details}")

                if room_notenough:
                    print(f"{room_notenough[0]}号室の詳細情報をbuilding_dataに反映します。")
                    # room_notenough[0]が該当する部屋番号の辞書を見つけて、room_detailsで更新する
                    for building in building_data:
                        if building_name in building:
                            for room in building[building_name]:
                                if room['部屋番号'] == str(room_notenough[0]):
                                    # room_detailsの値で更新
                                    for key in room_details:
                                        room[key] = room_details[key]
                    print("building_dataの更新が完了しました。")
                    # 結果を表示
                    print(building_data)

            room_notenough = []
            print("building_dataを再チェックし、いずれかの情報が空の部屋を探します。")
            # building_dataを参照して、いずれかの情報が空の部屋を探す
            for building in building_data:
                for building_name, rooms in building.items():
                    for room in rooms:
                        # いずれかのキーが空であればその部屋番号を追加
                        if any(room[key] == '' for key in ['坪数', '坪単価', '保証金', '入居時期']):
                            room_notenough.append(room['部屋番号'])
            print(f"情報が不完全な部屋番号:{room_notenough}")
            print(f"{building_number}番目のビルの空き部屋の部屋番号と詳細情報が全て収集されました。確認が必要なビル数は{len(building_names)}次のビルの空き部屋の部屋番号を尋ねます。")

            if room_notenough:
                print(f"情報が不完全な部屋番号が見つかりました: {room_notenough[0]}")
                first_room_number = room_notenough[0]
                missing_info = []
                print(f"{first_room_number}号室の不足している詳細情報を確認します。")
                for building in building_data:
                    for building_name, rooms in building.items():
                        for room in rooms:
                            if room['部屋番号'] == first_room_number:
                                # 不足しているキーをリストに追加
                                missing_info = [key for key in ['坪数', '坪単価', '保証金', '入居時期'] if room[key] == '']
                                break
                # 結果を出力
                print(f"情報が不完全な部屋番号:{room_notenough}\n一部屋目の不足情報:{missing_info}")
                # 空き部屋の詳細情報収集を続けるための応答生成
                stance=f"空き部屋の部屋番号は{room_notenough[0]}です。この部屋の{missing_info}を全て漏れの無いように簡潔に尋ねてください。"
                # 不完全な部屋番号を格納する変数

            elif building_number<len(building_names)-1:
                # 会話履歴を保存
                # conversation_txt_file_path=f"{conversation_history_file_path.split(".")[0]}_{company_name}_{building_name}.txt"
                save_conversation_history(conversation_history_file_path, conversation_history)
                move_file_to_date_folder(conversation_history_file_path, f"{company_name}_{building_name}")
                building_complete=building_names[building_number]
                print(f"{building_complete}の空き部屋の部屋番号と詳細情報が全て収集されました。")
                building_number+=1
                building_name=building_names[building_number]
                print(f"続いて、次のビルの空き部屋の部屋番号を尋ねます。ビル名は{building_name}です。")
                vacant_room_numbers = []
                conversation_history = []
                vacant_enough=False
                details_enough=False
                roomnumber_enough=False
                building_enough=False
                stance=f"続けて、次のビルの空き部屋の部屋番号を尋ねます。まず、ビル名{building_complete}について情報を提供してもらったことに対して感謝を述べ、続けて、ビル名{building_name}に空室があれば空室の部屋番号を教えてください、と相手に尋ねてください。"
    
            else:
                print("空き部屋の部屋番号と詳細情報が全て収集されたため、通話を終了します。")
                stance="空き部屋の部屋番号と詳細情報の収集を終えました。相手に感謝を告げ、自然に通話を終了するよう、最後の挨拶をしてください。"
                call_end=True

        # 会話履歴から応答文章を生成                    
        response_text = generate_response(stance, conversation_history)
        print(response_text)
        # 会話履歴に追加
        conversation_history.append({"role": "system", "content": response_text})
        print(f"会話履歴に追加しました：{conversation_history}")
        # 音声ファイルを生成
        text_to_speech(response_text, audio_file_path)
        print("音声ファイルを生成しました")
        # 音声ファイルを結合する
        concatenate_or_copy_recordings(output_file_path, audio_file_path, output_file_path)

        # 情報収集が不十分であれば通話を継続
        if call_end==False:
            # 音声ファイルを再生
            response = VoiceResponse()
            print("音声ファイルを再生します")
            response.play(public_url + '/serve_audio')
            print("音声ファイルを再生しました")
            # 再生が完了したら録音を開始するようにRedirect
            response.redirect('/start-recording')
            print("Redirect to /start-recording")
        
        # 情報収集が十分であれば通話を終了
        if call_end==True:
            # 会話履歴を保存
            # conversation_txt_file_path=f"{conversation_history_file_path.split(".")[0]}_{company_name}_{building_name}.txt"
            save_conversation_history(conversation_history_file_path, conversation_history)
            move_file_to_date_folder(conversation_history_file_path, f"{company_name}_{building_name}")
            conversation_history = []
            print(f"会話履歴を保存しました: {conversation_history_file_path}")
            # 音声ファイルを再生
            response = VoiceResponse()
            print("音声ファイルを再生します")
            response.play(public_url + '/serve_audio')
            print("音声ファイルを再生しました")
            # 終了
            # response.hangup()
            print("通話を終了します")

        return str(response)
        # return jsonify(message="録音と保存が完了しました"), 200

    except requests.exceptions.HTTPError as e:
        print(f"録音ファイルの取得に失敗しました: {e}")
        return jsonify(message="録音ファイルの取得に失敗しました"), 500

# 公開URLを使用して架電処理を開始する関数
def make_calls(public_url):
    global conversation_history
    global first_or
    global vacant_enough
    global call_end
    global building_data
    global building_name
    global details_enough
    global building_names
    global building_name
    global company_name
    global building_number
    global roomnumber_enough
    global room_notenough
    global missing_info
    global room_details
    global vacant_room_numbers
    global building_enough
    global room_details
    global building_number
    global roomnumber_enough
    global room_notenough
    global missing_info
    global vacant_enough
    global details_enough
    
    building_data = []

    print(f"公開URLを使用して架電処理を開始します: {public_url}")
    
    # 管理会社・オーナー名ごとにグループ化
    grouped_df = df.groupby('管理会社・オーナー名')
    print(f"グループ化されたデータフレーム: {grouped_df}")

    for company_name, group in grouped_df:
        vacant_enough = False
        details_enough = False
        call_end = False
        first_or = True  # 初期値としてTrueを設定
        building_number=0
        vacant_room_numbers = []
        conversation_history = []
        vacant_enough=False
        details_enough=False
        roomnumber_enough=False
        building_enough=False
        vacant_rooms = []
        room_notenough = []
        missing_info = []
        room_details = []
        conversation_history = []
        
        # 架電対象の電話番号とビル名を取得
        phone_number = group['電話番号'].iloc[0]
        building_codes = group['ビルコード（7ケタ）'].tolist()
        building_names = group['ビル名称'].tolist()
        print(f"電話番号を確認します: {phone_number}")
        print(f"ビルコードを確認します: {building_codes}")
        print(f"ビル名を確認します: {building_names}")
        building_name=building_names[building_number]
        print(f"最初に空室確認を行うビル: {building_name}")
       
        # 電話番号がNaNでないことを確認
        if pd.isna(phone_number):
            print("電話番号がNaNのためスキップします")
            continue

        # 電話番号を変換
        converted_phone_number = convert_phone_number(phone_number)
        print(f"変換後の電話番号: {converted_phone_number}に発信します")
        
        # 挨拶文の作成（複数のビル名を含める）
        stance_greeting = f'あなたは不動産仲介業者の{my_info}です。あなたは、{company_name}に電話をかけて、自分の会社名と個人名を名乗ったうえで、電話相手の不動産管理会社の担当者に極めて簡単な挨拶をし、相手の会社が管理している不動産の各部屋の空き状況を聞き出したい。空室の情報を知りたいビルは、{building_names}の全てです。一棟ずつ空室情報を確認していくために、まずは{building_name}について空室情報を聞く旨を伝え、当該ビルで現在空室となっている部屋番号を聞いてくれ。'
        conversation_history = []  # 例として空のリストを使用
        model = "gpt-4o-mini"
        greeting_text = generate_response(stance_greeting, conversation_history, model)
        print(f"生成された挨拶文: {greeting_text}")

        # 会話履歴に追加
        conversation_history.append({"role": "system", "content": greeting_text})
        print(f"会話履歴に追加しました：{conversation_history}")        
        # 挨拶を音声ファイルに変換
        text_to_speech(greeting_text, audio_file_path)
        print("音声ファイルに変換しました")

        # 音声ファイルをコピー
        concatenate_or_copy_recordings("", audio_file_path, output_file_path)
        print("音声ファイルをコピーしました")

        # Twilioを使用して通話を開始し、挨拶を再生
        call = client.calls.create(
            url=f'{public_url}/twiml',
            to=converted_phone_number,
            from_=from_number
        )
        print(f"Call SID: {call.sid} to {phone_number}")

        while True:
            call_status = client.calls(call.sid).fetch().status
            # print(f"Call status: {call_status}")
            if call_status in ['completed', 'failed', 'busy', 'no-answer']:
                print(f"Call to {converted_phone_number} ended with status: {call_status}")
                break
            time.sleep(1)

        # # openpyxlで空室確認を行ったビルに対応するシートのC6セルに今日の日付を格納
        # for building_code in building_codes:
        #     try:
        #         if building_code in wb.sheetnames:
        #             sheet = wb[building_code]
        #             sheet['C6'] = current_date
        #             print(f"Updated C6 in sheet {building_code} with {current_date}")
        #         else:
        #             print(f"Sheet for building code {building_code} not found")
        #     except Exception as e:
        #         print(f"Error updating sheet {building_code}: {e}")

        # # 会話履歴を保存
        # save_conversation_history(conversation_history_file_path, conversation_history)
        # move_file_to_date_folder(conversation_history_file_path, company_name+building_name)
        # print(f"会話履歴を保存しました: {conversation_txt_file_path}")

        # 音声ファイルをコピー
        move_file_to_date_folder(output_file_path, company_name)
        print("通話音声ファイルを作成しました")

        print("全ての架電処理が完了しました")
        
    # ビルデータ内情報を部屋番号でソート
    sorted_building_data = sort_vacant_rooms_in_building(building_data)
    print(f"ソート後のビルデータ: {sorted_building_data}")

    # エクセルファイルから空室情報を削除
    remove_vacant_columns(df)

    # 関数を呼び出して、H列以降に空室情報を挿入
    df_updated = insert_vacant_info_to_columns(sorted_building_data, df)

    # 更新したデータをExcelファイルに保存
    df_updated.to_excel(copied_file_path, index=False, engine='openpyxl')
    print(f"更新後のExcelファイルを保存しました: {copied_file_path}")    

# Flaskサーバーを起動する関数
def start_flask_server():
    print("Flaskサーバーを起動します")
    app.run(port=5000)

# メイン関数
if __name__ == '__main__':
    try:
        port = 5000
        print("ngrokトンネルを確立します")
        public_url = ngrok.connect(port).public_url
        print(f"ngrokトンネルが確立されました: '{public_url}' -> 'http://127.0.0.1:{port}'")

        print("Flaskサーバーを別スレッドで起動します")
        flask_thread = threading.Thread(target=start_flask_server)
        flask_thread.daemon = True  # これでメインスレッド終了時にFlaskサーバーも終了
        flask_thread.start()

        print("架電処理を開始します")
        make_calls(public_url)

    except KeyboardInterrupt:
        print("KeyboardInterruptが検出されました。終了します...")


