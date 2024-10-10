import pandas as pd
from openpyxl import load_workbook
from twilio.rest import Client
from dotenv import load_dotenv
import os
import time
from flask import Flask, request, jsonify, send_file
from twilio.twiml.voice_response import VoiceResponse
import threading
from pyngrok import ngrok
from pathlib import Path
import openai
import requests
from datetime import datetime
import shutil
from pydub import AudioSegment
from queue import Queue
import re
import ast
import functions as fc

# Flaskアプリケーションの作成
app = Flask(__name__)

# ファイルから環境変数を読み込む
print("環境変数を.envファイルから読み込みます")
load_dotenv()

# グローバル変数
audio_file_path = 'audio.mp3'
output_file_path = 'output.mp3'
account_sid = os.getenv('TWILIO_SID')
auth_token = os.getenv('TWILIO_AUTH_TOKEN')
from_number = os.getenv('TWILIO_PHONE_NUMBER')
my_info="株式会社テスト、業務改革室の竹内です。"
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
first_or = True  # 初期値としてTrueを設定
vacant_enough=False
call_end=False
roomnumber_enough=False
building_enough=False
conversation_history=[]
room_details=[]
vacant_rooms=[]
building_data = []
missing_info=[]
room_notenough=[]

# 現在の日付を取得
current_date = datetime.now().strftime('%Y/%m/%d')

# Twilioの設定
print("Twilioクライアントを初期化します")
client = Client(account_sid, auth_token)
print("Twilioクライアントが初期化されました")

# Excelファイルのパスを設定
print("Excelファイルのパスを設定します")
excel_file_path = 'C:\\Users\\ntake\\my_project\\bukkaku_call_automation\\デモ用ビルデータサンプル.xlsx'
print(f"Excelファイルパス: {excel_file_path}")

# コピーするExcelファイルのパス
copied_file_path = f'{excel_file_path.split(".")[0]}_copy.xlsx'
shutil.copyfile(excel_file_path, copied_file_path)

# 会話履歴のファイルパスを設定
conversation_history_file_path = 'C:\\Users\\ntake\\my_project\\bukkaku_call_automation\\conversation_history.txt'

# Excelファイルを読み込み
print("Excelファイルを読み込みます")
df = pd.read_excel(copied_file_path, dtype={'電話番号': str, 'ビルコード（7ケタ）': str,"ビル名称":str})

# openpyxlを使用して同じExcelファイルを開く
wb = load_workbook(copied_file_path)

# Twilioの環境変数が正しく設定されているか確認
if not account_sid or not auth_token or not from_number:
    print("Twilioの環境変数が正しく設定されていません。")
    raise ValueError("Twilioの環境変数が正しく設定されていません。")

# TwiMLエンドポイント(メインのエンドポイント)
@app.route('/twiml', methods=['POST'])
def twiml():
    print("TwiMLエンドポイントが呼び出されました")
    
    # TwiMLレスポンスを生成
    response = VoiceResponse()
    
    # 挨拶用の音声ファイルを再生
    response.play(public_url + '/serve_audio')  # 音声ファイルのURL
    print("音声ファイルを再生しました")
    
    # 再生が完了したら録音を開始するようにRedirect
    response.redirect('/start-recording')
    print("Redirect to /start-recording")
    
    print("TwiMLレスポンスが生成されました")

    return str(response)

# Flaskルート: システム発話用の音声ファイルを提供するためのエンドポイント
@app.route('/serve_audio')
def serve_audio():
    print("システム発話用の音声ファイルを提供しています")
    try:
        print("システム発話用の音声ファイルを提供しています")
        return send_file(audio_file_path, mimetype='audio/mpeg')
    except Exception as e:
        print(f"システム発話用の音声ファイルの提供中にエラーが発生しました: {e}")
        return jsonify(error=str(e)), 500

# 録音を開始するエンドポイント
@app.route('/start-recording', methods=['POST'])
def start_recording():
    print("録音を開始します")
    response = VoiceResponse()
    response.record(max_length=60, action='/handle-recording', method="POST", transcribe=False)
    print("録音を開始しました")
    return str(response)

# 録音が完了した後のエンドポイント
@app.route('/handle-recording', methods=['POST'])
def handle_recording():

    global conversation_history
    global vacant_enough
    global vacant_room_numbers
    global details_enough
    global building_data
    global building_name
    global missing_info
    global room_notenough
    global company_name
    global building_names
    global room_details
    global building_number
    global roomnumber_enough
    global building_enough
    call_end=False
        
    print("録音処理を開始します")
    
    recording_url = request.form['RecordingUrl']
    print(f"録音URLを受信しました: {recording_url}")
   
    time.sleep(3)  # 3秒待機
    print("3秒待機しました。録音ファイルをダウンロードします。")

    try:
        response = requests.get(recording_url, auth=(account_sid, auth_token))
        response.raise_for_status()
        print("録音ファイルのダウンロードが成功しました。")

        with open(audio_file_path, 'wb') as audio_file:
            audio_file.write(response.content)
            print(f"録音ファイルを保存しました: {audio_file_path}")
        # 音声ファイルを結合        
        fc.concatenate_or_copy_recordings(output_file_path, audio_file_path, output_file_path)

        # 録音ファイルをテキスト化
        transcription = fc.transcribe_audio(audio_file_path)
        print(f"テキスト化された内容: {transcription}")

        # 会話履歴に追加
        conversation_history.append({"role": "user", "content": transcription})
        print(f"会話履歴に追加しました：{conversation_history}")
        # conversation_txt_file_path=f"{conversation_history_file_path.split(".")[0]}_{company_name}_{building_name}.txt"
        
        # 対象ビルの空き部屋の部屋番号が十分に収集されているかどうかを判定
        if building_enough!=True:
            building_enough=fc.judgement_enough(conversation_history_file_path, conversation_history)

        # 対象ビルの空き部屋の部屋番号が十分に収集されている場合
        if building_enough==True and vacant_enough==True and roomnumber_enough!=True:
            # 会話履歴から空き部屋の部屋番号を抽出
            print(f"[INFO] Vacant rooms before processing: {conversation_history}")
            vacant_room_numbers = fc.extract_room_numbers(conversation_history)
            print(f"[INFO] Extracted vacant room numbers: {vacant_room_numbers}")
            print("空き部屋の部屋番号が十分に収集されています")

            # ビルデータ（部屋番号のみ）の作成
            building_data = fc.add_vacant_rooms_to_building(building_data,building_name, vacant_room_numbers)
            print(f"[INFO] Updated building data: {building_data}")
            print(building_data)
            print(f"ビルデータに空室情報を追加しました: {building_data}")
            # ビルデータ内の情報を部屋番号でソート
            building_data = fc.sort_vacant_rooms_in_building(building_data)

            roomnumber_enough=True
        
        # 対象ビルの空き部屋の部屋番号が十分に収集されていない場合
        elif roomnumber_enough!=True:
            # 会話履歴から空き部屋の部屋番号を抽出
            print(f"[INFO] Vacant rooms before processing: {conversation_history}")
            vacant_room_numbers = fc.extract_room_numbers(conversation_history)
            print(f"[INFO] Extracted vacant room numbers: {vacant_room_numbers}")
            print("空き部屋の部屋番号が不十分です")
            # 空き部屋の部屋番号収集を続けるための応答生成
            stance=f"現状確認できている空き部屋の部屋番号は以下の通りです。{vacant_room_numbers}ここで空き確認が取れている部屋番号を相手に伝え、他にも空いている部屋が無いか尋ねてください。"
            vacant_enough=True

        # 各部屋の詳細情報収集
        if roomnumber_enough==True:
            print("部屋番号の情報が十分に収集されたため、部屋の詳細情報を抽出します。")
            
            if missing_info:
                room_details=fc.extract_room_details(missing_info,conversation_history)
                room_details=fc.dict_from_str(room_details)
                print(f"部屋の詳細情報:{room_details}")

                if room_notenough:
                    print(f"{room_notenough[0]}号室の詳細情報をbuilding_dataに反映します。")
                    # room_notenough[0]が該当する部屋番号の辞書を見つけて、room_detailsで更新する
                    for building in building_data:
                        if building_name in building:
                            for room in building[building_name]:
                                if room['部屋番号'] == str(room_notenough[0]):
                                    # room_detailsの値で更新
                                    for key in room_details:
                                        room[key] = room_details[key]
                    print("building_dataの更新が完了しました。")
                    # 結果を表示
                    print(building_data)

            room_notenough = []
            print("building_dataを再チェックし、いずれかの情報が空の部屋を探します。")
            # building_dataを参照して、いずれかの情報が空の部屋を探す
            for building in building_data:
                for building_name, rooms in building.items():
                    for room in rooms:
                        # いずれかのキーが空であればその部屋番号を追加
                        if any(room[key] == '' for key in ['坪数', '坪単価', '保証金', '入居時期']):
                            room_notenough.append(room['部屋番号'])
            print(f"情報が不完全な部屋番号:{room_notenough}")
            print(f"{building_number}番目のビルの空き部屋の部屋番号と詳細情報が全て収集されました。確認が必要なビル数は{len(building_names)}次のビルの空き部屋の部屋番号を尋ねます。")

            if room_notenough:
                print(f"情報が不完全な部屋番号が見つかりました: {room_notenough[0]}")
                first_room_number = room_notenough[0]
                missing_info = []
                print(f"{first_room_number}号室の不足している詳細情報を確認します。")
                for building in building_data:
                    for building_name, rooms in building.items():
                        for room in rooms:
                            if room['部屋番号'] == first_room_number:
                                # 不足しているキーをリストに追加
                                missing_info = [key for key in ['坪数', '坪単価', '保証金', '入居時期'] if room[key] == '']
                                break
                # 結果を出力
                print(f"情報が不完全な部屋番号:{room_notenough}\n一部屋目の不足情報:{missing_info}")
                # 空き部屋の詳細情報収集を続けるための応答生成
                stance=f"空き部屋の部屋番号は{room_notenough[0]}です。この部屋の{missing_info}を全て漏れの無いように簡潔に尋ねてください。"
                # 不完全な部屋番号を格納する変数

            elif building_number<len(building_names)-1:
                # 会話履歴を保存
                # conversation_txt_file_path=f"{conversation_history_file_path.split(".")[0]}_{company_name}_{building_name}.txt"
                fc.save_conversation_history(conversation_history_file_path, conversation_history)
                fc.move_file_to_date_folder(company_name, conversation_history_file_path, f"{company_name}_{building_name}")
                building_complete=building_names[building_number]
                print(f"{building_complete}の空き部屋の部屋番号と詳細情報が全て収集されました。")
                building_number+=1
                building_name=building_names[building_number]
                print(f"続いて、次のビルの空き部屋の部屋番号を尋ねます。ビル名は{building_name}です。")
                vacant_room_numbers = []
                conversation_history = []
                vacant_enough=False
                details_enough=False
                roomnumber_enough=False
                building_enough=False
                stance=f"続けて、次のビルの空き部屋の部屋番号を尋ねます。まず、ビル名{building_complete}について情報を提供してもらったことに対して感謝を述べ、続けて、ビル名{building_name}に空室があれば空室の部屋番号を教えてください、と相手に尋ねてください。"
    
            else:
                print("空き部屋の部屋番号と詳細情報が全て収集されたため、通話を終了します。")
                stance="空き部屋の部屋番号と詳細情報の収集を終えました。相手に感謝を告げ、自然に通話を終了するよう、最後の挨拶をしてください。"
                call_end=True

        # 会話履歴から応答文章を生成                    
        response_text = fc.generate_response(stance, conversation_history)
        print(response_text)
        # 会話履歴に追加
        conversation_history.append({"role": "system", "content": response_text})
        print(f"会話履歴に追加しました：{conversation_history}")
        # 音声ファイルを生成
        fc.text_to_speech(response_text, audio_file_path)
        print("音声ファイルを生成しました")
        # 音声ファイルを結合する
        fc.concatenate_or_copy_recordings(output_file_path, audio_file_path, output_file_path)

        # 情報収集が不十分であれば通話を継続
        if call_end==False:
            # 音声ファイルを再生
            response = VoiceResponse()
            print("音声ファイルを再生します")
            response.play(public_url + '/serve_audio')
            print("音声ファイルを再生しました")
            # 再生が完了したら録音を開始するようにRedirect
            response.redirect('/start-recording')
            print("Redirect to /start-recording")
        
        # 情報収集が十分であれば通話を終了
        if call_end==True:
            # 会話履歴を保存
            # conversation_txt_file_path=f"{conversation_history_file_path.split(".")[0]}_{company_name}_{building_name}.txt"
            fc.save_conversation_history(conversation_history_file_path, conversation_history)
            fc.move_file_to_date_folder(company_name, conversation_history_file_path, f"{company_name}_{building_name}")
            conversation_history = []
            print(f"会話履歴を保存しました: {conversation_history_file_path}")
            # 音声ファイルを再生
            response = VoiceResponse()
            print("音声ファイルを再生します")
            response.play(public_url + '/serve_audio')
            print("音声ファイルを再生しました")
            # 終了
            # response.hangup()
            print("通話を終了します")

        return str(response)
        # return jsonify(message="録音と保存が完了しました"), 200

    except requests.exceptions.HTTPError as e:
        print(f"録音ファイルの取得に失敗しました: {e}")
        return jsonify(message="録音ファイルの取得に失敗しました"), 500

# 公開URLを使用して架電処理を開始する関数
def make_calls(public_url):
    global conversation_history
    global first_or
    global vacant_enough
    global call_end
    global building_data
    global building_name
    global details_enough
    global building_names
    global building_name
    global company_name
    global building_number
    global roomnumber_enough
    global room_notenough
    global missing_info
    global room_details
    global vacant_room_numbers
    global building_enough
    global room_details
    global building_number
    global roomnumber_enough
    global room_notenough
    global missing_info
    global vacant_enough
    global details_enough
    
    building_data = []

    print(f"公開URLを使用して架電処理を開始します: {public_url}")
    
    # 管理会社・オーナー名ごとにグループ化
    grouped_df = df.groupby('管理会社・オーナー名')
    print(f"グループ化されたデータフレーム: {grouped_df}")

    for company_name, group in grouped_df:
        vacant_enough = False
        details_enough = False
        call_end = False
        first_or = True  # 初期値としてTrueを設定
        building_number=0
        vacant_room_numbers = []
        conversation_history = []
        vacant_enough=False
        details_enough=False
        roomnumber_enough=False
        building_enough=False
        vacant_rooms = []
        room_notenough = []
        missing_info = []
        room_details = []
        conversation_history = []
        
        # 架電対象の電話番号とビル名を取得
        phone_number = group['電話番号'].iloc[0]
        building_codes = group['ビルコード（7ケタ）'].tolist()
        building_names = group['ビル名称'].tolist()
        print(f"電話番号を確認します: {phone_number}")
        print(f"ビルコードを確認します: {building_codes}")
        print(f"ビル名を確認します: {building_names}")
        building_name=building_names[building_number]
        print(f"最初に空室確認を行うビル: {building_name}")
       
        # 電話番号がNaNでないことを確認
        if pd.isna(phone_number):
            print("電話番号がNaNのためスキップします")
            continue

        # 電話番号を変換
        converted_phone_number = fc.convert_phone_number(phone_number)
        print(f"変換後の電話番号: {converted_phone_number}に発信します")
        
        # 挨拶文の作成（複数のビル名を含める）
        stance_greeting = f'あなたは不動産仲介業者の{my_info}です。あなたは、{company_name}に電話をかけて、自分の会社名と個人名を名乗ったうえで、電話相手の不動産管理会社の担当者に極めて簡単な挨拶をし、相手の会社が管理している不動産の各部屋の空き状況を聞き出したい。空室の情報を知りたいビルは、{building_names}の全てです。一棟ずつ空室情報を確認していくために、まずは{building_name}について空室情報を聞く旨を伝え、当該ビルで現在空室となっている部屋番号を聞いてくれ。'
        conversation_history = []  # 例として空のリストを使用
        model = "gpt-4o-mini"
        greeting_text = fc.generate_response(stance_greeting, conversation_history, model)
        print(f"生成された挨拶文: {greeting_text}")

        # 会話履歴に追加
        conversation_history.append({"role": "system", "content": greeting_text})
        print(f"会話履歴に追加しました：{conversation_history}")        
        # 挨拶を音声ファイルに変換
        fc.text_to_speech(greeting_text, audio_file_path)
        print("音声ファイルに変換しました")

        # 音声ファイルをコピー
        fc.concatenate_or_copy_recordings("", audio_file_path, output_file_path)
        print("音声ファイルをコピーしました")

        # Twilioを使用して通話を開始し、挨拶を再生
        call = client.calls.create(
            url=f'{public_url}/twiml',
            to=converted_phone_number,
            from_=from_number
        )
        print(f"Call SID: {call.sid} to {phone_number}")

        while True:
            call_status = client.calls(call.sid).fetch().status
            # print(f"Call status: {call_status}")
            if call_status in ['completed', 'failed', 'busy', 'no-answer']:
                print(f"Call to {converted_phone_number} ended with status: {call_status}")
                break
            time.sleep(1)

        # # openpyxlで空室確認を行ったビルに対応するシートのC6セルに今日の日付を格納
        # for building_code in building_codes:
        #     try:
        #         if building_code in wb.sheetnames:
        #             sheet = wb[building_code]
        #             sheet['C6'] = current_date
        #             print(f"Updated C6 in sheet {building_code} with {current_date}")
        #         else:
        #             print(f"Sheet for building code {building_code} not found")
        #     except Exception as e:
        #         print(f"Error updating sheet {building_code}: {e}")

        # # 会話履歴を保存
        # save_conversation_history(conversation_history_file_path, conversation_history)
        # move_file_to_date_folder(company_name, conversation_history_file_path, company_name+building_name)
        # print(f"会話履歴を保存しました: {conversation_txt_file_path}")

        # 音声ファイルをコピー
        fc.move_file_to_date_folder(company_name, output_file_path, company_name)
        print("通話音声ファイルを作成しました")

        print("全ての架電処理が完了しました")
        
    # ビルデータ内情報を部屋番号でソート
    sorted_building_data = fc.sort_vacant_rooms_in_building(building_data)
    print(f"ソート後のビルデータ: {sorted_building_data}")

    # エクセルファイルから空室情報を削除
    fc.remove_vacant_columns(df)

    # 関数を呼び出して、H列以降に空室情報を挿入
    df_updated = fc.insert_vacant_info_to_columns(sorted_building_data, df)

    # 更新したデータをExcelファイルに保存
    df_updated.to_excel(copied_file_path, index=False, engine='openpyxl')
    print(f"更新後のExcelファイルを保存しました: {copied_file_path}")    

# Flaskサーバーを起動する関数
def start_flask_server():
    print("Flaskサーバーを起動します")
    app.run(port=5000)

# メイン関数
if __name__ == '__main__':
    try:
        port = 5000
        print("ngrokトンネルを確立します")
        public_url = ngrok.connect(port).public_url
        print(f"ngrokトンネルが確立されました: '{public_url}' -> 'http://127.0.0.1:{port}'")

        print("Flaskサーバーを別スレッドで起動します")
        flask_thread = threading.Thread(target=start_flask_server)
        flask_thread.daemon = True  # これでメインスレッド終了時にFlaskサーバーも終了
        flask_thread.start()

        print("架電処理を開始します")
        make_calls(public_url)

    except KeyboardInterrupt:
        print("KeyboardInterruptが検出されました。終了します...")


